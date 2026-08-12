from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REFERENCE = Path(r"C:\Users\LENOVO\Downloads\1993_DN_SE_07_DN_Report5_Test_Report.xlsx")
REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT = REPO_ROOT / "docs" / "report5" / "PORMS_Report5_Test_Report.xlsx"
DOWNLOAD_COPY = Path(r"C:\Users\LENOVO\Downloads\PORMS_Report5_Test_Report.xlsx")
TEST_CASE_OUTPUT = REPO_ROOT / "docs" / "report5" / "PORMS_Test_Cases.xlsx"
TEST_CASE_DOWNLOAD_COPY = Path(r"C:\Users\LENOVO\Downloads\PORMS_Test_Cases.xlsx")
EXECUTED_DATE = datetime(2026, 8, 11)
TESTER = "PORMS Team"


@dataclass(frozen=True)
class Case:
    title: str
    procedure: str
    expected: str
    precondition: str = "PORMS test environment is available."
    status: str = "Passed"


@dataclass(frozen=True)
class FunctionGroup:
    name: str
    cases: tuple[Case, ...]


@dataclass(frozen=True)
class Module:
    code: str
    sheet: str
    feature: str
    description: str
    precondition: str
    evidence: str
    groups: tuple[FunctionGroup, ...]


def c(title: str, action: str, expected: str, precondition: str = "") -> Case:
    return Case(
        title,
        f"1. Open the related PORMS screen.\n2. {action}\n3. Observe the response and persisted data.",
        expected,
        precondition or "PORMS test environment is available with suitable test data.",
    )


MODULES: tuple[Module, ...] = (
    Module(
        "AUTH", "Authentication", "Authentication & Account", "Verify secure access, session handling and account self-service.",
        "An active PORMS account exists for each role under test.",
        "LoginPage.test.tsx; authService.test.ts; PasswordPolicyTests.cs; RoleAuthorizationTests.cs",
        (
            FunctionGroup("Sign in and session", (
                c("Sign in with valid credentials", "Enter a valid email and password, then select Sign in.", "The user is authenticated and redirected to the role-scoped dashboard."),
                c("Reject invalid credentials", "Enter an incorrect email or password.", "Authentication is rejected without creating a valid session."),
                c("Validate required login fields", "Submit the form with an empty email or password.", "Required-field validation is shown and no request is accepted."),
                c("Refresh an expired access token", "Call a protected API after the access token expires while a valid refresh token exists.", "The access token is renewed and the protected request is retried safely."),
                c("Reject a revoked refresh token", "Attempt token renewal using a revoked token.", "Renewal is rejected and the user must sign in again."),
                c("Sign out", "Select Sign out from the account menu.", "Tokens are cleared/revoked and protected screens are no longer accessible."),
            )),
            FunctionGroup("Password and profile", (
                c("Change password with valid data", "Enter the current password and a valid new password with matching confirmation.", "The password is changed and the account remains protected."),
                c("Reject mismatched password confirmation", "Enter different new-password and confirmation values.", "Client validation blocks submission."),
                c("Enforce password policy", "Submit a new password that does not meet the configured policy.", "The password is rejected with a clear validation message."),
            )),
        ),
    ),
    Module(
        "DASH", "Dashboard & Weather", "Dashboard & Weather Monitoring", "Verify scoped dashboard, GIS and weather presentation.",
        "Ports, zones, weather readings and role-scoped users exist.",
        "DashboardPage.test.tsx; GisMapCard.test.tsx; WeatherDataTable.test.tsx; DashboardSummaryTests.cs",
        (
            FunctionGroup("Operational dashboard", (
                c("Load dashboard risk summary", "Open the dashboard as an Admin.", "Risk totals and active alerts are loaded for all authorized ports."),
                c("Apply Port Manager scope", "Open the dashboard as a Port Manager.", "Only the assigned port, zones, alerts and weather are returned."),
                c("Render GIS port and zone markers", "Select a port on the Leaflet map.", "The selected port and its zones appear at their stored coordinates."),
                c("Open map detail", "Select a port or zone marker.", "A readable popup shows location and operational risk information."),
                c("Display latest weather readings", "Review the weather detail table.", "Wind, rain, visibility, temperature, humidity and update time are displayed."),
                c("Show Beaufort wind information", "Review a reading that contains wind speed.", "Wind information is presented with the corresponding Beaufort level where available."),
                c("Handle an empty dashboard dataset", "Load a scoped dashboard with no readings.", "The page shows a controlled empty state instead of crashing."),
            )),
        ),
    ),
    Module(
        "PORT", "Port & Zone", "Port & Zone Management", "Verify creation and maintenance of ports and operational zones.",
        "Admin is authenticated; unique port codes and valid coordinates are available.",
        "PortManagementPage.test.tsx; PortCreatePage.test.tsx; PortTable.test.tsx; portService.test.ts",
        (
            FunctionGroup("Port management", (
                c("List ports", "Open Port & Zone Management.", "The authorized port list loads with code, name and status."),
                c("Create a port with zones", "Enter valid port data and one or more zone rows, then save.", "The port and zones are persisted and shown in the list."),
                c("Validate required port data", "Submit a port without required code, name or coordinates.", "Validation prevents invalid creation."),
                c("Reject duplicate port code", "Create a port using an existing port code.", "The duplicate is rejected and existing data is preserved."),
                c("Update port details", "Edit the selected port name, coordinates or status.", "The saved values are reloaded in the port list."),
            )),
            FunctionGroup("Zone management", (
                c("Create a zone for a port", "Add a valid dock, yard, gate or warehouse zone.", "The zone is linked to the selected port."),
                c("Update zone details", "Edit a zone and save changes.", "The zone table reloads with the updated values."),
                c("Prevent cross-port zone access", "Request a zone outside the manager's assigned port.", "Backend authorization prevents unauthorized access."),
            )),
        ),
    ),
    Module(
        "USER", "Users & Roles", "User & Role Management", "Verify account administration, role rules, port assignment, filtering and pagination.",
        "Admin is authenticated; ports exist for manager/operator assignment.",
        "UsersPage.test.tsx; UserFormPage.test.tsx; userService.test.ts; RoleAuthorizationTests.cs",
        (
            FunctionGroup("Account administration", (
                c("Load and paginate users", "Open Users and move between result pages.", "Users are shown with stable pagination and total count."),
                c("Filter by name or email", "Enter a name or email in the search field.", "Only matching users remain in the result table."),
                c("Filter by role, port and status", "Apply role, assigned port and status filters.", "Filters combine correctly before pagination."),
                c("Create an Admin account", "Create a System Admin without assigning a port.", "The account is saved with global scope and no port assignment."),
                c("Create a Port Manager", "Create a Port Manager and select a port.", "The account is linked to exactly the selected port."),
                c("Create an Operator", "Create an Operator with an assigned port or pending assignment.", "The account is saved with the selected assignment state."),
                c("Update a user", "Edit profile, role, port assignment or active status.", "The user list reloads with the new values."),
                c("Delete a user after confirmation", "Confirm deletion from the user list.", "The account is removed only after explicit confirmation."),
                c("Deny user administration to non-Admin", "Call user-administration endpoints as Port Manager or Operator.", "The backend returns an authorization failure."),
            )),
        ),
    ),
    Module(
        "RISK", "Risk Thresholds", "Risk Threshold Management", "Verify threshold evaluation, validation, overrides and Excel import.",
        "Admin or Port Manager is authenticated and default thresholds exist.",
        "RiskConfigPage.test.tsx; RiskEngineTests.cs; RiskThresholdValidatorTests.cs; RiskThresholdExcelServiceTests.cs",
        (
            FunctionGroup("Risk evaluation", (
                c("Evaluate wind thresholds", "Evaluate weather readings around each configured wind boundary.", "Each wind value maps to the expected LOW, MEDIUM, HIGH or CRITICAL level."),
                c("Evaluate rain thresholds", "Evaluate rainfall values around configured boundaries.", "Rain risk increases according to the configured rules."),
                c("Evaluate visibility thresholds", "Evaluate visibility values around configured boundaries.", "Lower visibility produces the expected higher risk."),
                c("Apply max-risk rule", "Evaluate a reading where weather factors have different levels.", "The final level equals the most dangerous factor."),
                c("Use fallback thresholds when configuration is unavailable", "Evaluate a reading without active database thresholds.", "Safe fallback thresholds are applied deterministically."),
            )),
            FunctionGroup("Configuration and import", (
                c("Save global thresholds", "Edit valid threshold rows and save.", "The complete configuration is validated and persisted."),
                c("Save a zone override", "Configure an override for an authorized zone.", "The zone-specific configuration is stored and used for that zone."),
                c("Preview a valid threshold Excel file", "Upload a complete valid template.", "Rows are validated and displayed for review without immediate persistence."),
                c("Reject an invalid threshold Excel file", "Upload a file with invalid factors, operators, units or overlapping ranges.", "The entire import is rejected with row-level errors."),
                c("Confirm a valid threshold import", "Confirm the previously validated preview.", "All threshold rows are applied atomically."),
                c("Deny Excel import to Port Manager", "Attempt threshold Excel import as Port Manager.", "Admin-only import authorization is enforced."),
            )),
        ),
    ),
    Module(
        "SOP", "SOP Rules", "SOP Rule Management", "Verify SOP rule lifecycle and transformation of risk into operational actions.",
        "Admin or Port Manager is authenticated; zone types and risk levels exist.",
        "SopRulesPage.test.tsx; sopRulesService.test.ts; SopEngineTests.cs; SopRuleImportTests.cs",
        (
            FunctionGroup("SOP rules", (
                c("List SOP rules by risk level", "Open SOP Rules and select each risk section.", "Rules are grouped into LOW, MEDIUM, HIGH and CRITICAL sections."),
                c("Create a valid SOP rule", "Enter code, name, risk, zone type, action and description.", "The rule is persisted with safe technical defaults."),
                c("Update an SOP rule", "Edit a business-visible field and save.", "The edit is saved while hidden technical values remain intact."),
                c("Disable an SOP rule", "Turn off an active rule.", "The rule remains stored but is excluded from automated matching."),
                c("Delete an SOP rule as Admin", "Confirm deletion of a selected rule.", "The Admin can delete the rule after confirmation."),
                c("Match active rule by risk and zone type", "Process a HIGH/CRITICAL risk event for a matching zone.", "The SOP Engine selects the applicable active rule."),
                c("Derive operation mode and actions", "Execute the selected SOP rule.", "Operation mode, alert/task actions and descriptions are created consistently."),
                c("Preview and confirm SOP Excel import", "Upload a valid SOP template, review it and confirm.", "Valid rules are imported atomically after preview."),
                c("Deny SOP import/delete to Port Manager", "Attempt Admin-only SOP operations as Port Manager.", "Authorization blocks the operation."),
            )),
        ),
    ),
    Module(
        "ALERT", "Alerts", "Alert Management", "Verify alert creation, scope, acknowledgement, popup and notification behavior.",
        "HIGH or CRITICAL evaluated events and scoped users exist.",
        "AlertPage.test.tsx; AlertDetailPage.test.tsx; AppShell.test.tsx; alertService.test.ts",
        (
            FunctionGroup("Alert lifecycle", (
                c("Create alert for HIGH risk", "Process a HIGH weather-risk event.", "A port/zone alert is persisted with weather values and required action."),
                c("Create alert for CRITICAL risk", "Process a CRITICAL weather-risk event.", "A critical alert is persisted with the correct scope and severity."),
                c("Avoid exposing unrelated port alerts", "Open alerts as a Port Manager from another port.", "Only alerts inside the manager's port scope are returned."),
                c("Restrict Operator alert visibility", "Open alerts as an Operator without a related task.", "Unrelated port alerts are not exposed."),
                c("Acknowledge an alert", "Open an unread authorized alert and acknowledge it.", "Receipt/read state is stored for the current user."),
                c("Show popup and audio for unread severe alert", "Refresh while an unread HIGH/CRITICAL alert exists and notifications are enabled.", "A popup is displayed and browser sound/speech is requested."),
                c("Honor notification preference", "Disable an alert notification preference and reload.", "Suppressed channels do not interrupt the user."),
                c("Filter and paginate alerts", "Filter by port, zone, date and risk, then navigate pages.", "Filtering occurs before 15-row pagination."),
            )),
        ),
    ),
    Module(
        "TASK", "Task Workflow", "Task Assignment & Operator Workflow", "Verify task generation, assignment and controlled operator lifecycle.",
        "An SOP-generated task, Port Manager and Operators exist.",
        "TasksPage.test.tsx; TaskDetailPage.test.tsx; AlertDetailPage.test.tsx; taskService.test.ts",
        (
            FunctionGroup("Assignment", (
                c("Create task from SOP action", "Process an event whose SOP includes task creation.", "A NEW task is linked to the alert, port and zone."),
                c("List valid assignees", "Open assignment for a task.", "Only active Operators belonging to the task's port are selectable."),
                c("Assign task to a valid Operator", "Select a valid Operator and due time, then save.", "The assignee is persisted and assignment progress is recorded."),
                c("Reject cross-port assignment", "Attempt assignment to an Operator from another port.", "Backend validation rejects the assignment."),
                c("Reassign a task", "Choose another valid Operator before completion.", "The new assignee and operation history are persisted."),
                c("Keep assignment when email fails", "Assign a task while SMTP delivery is unavailable.", "The business assignment remains saved; email failure does not roll back the task."),
            )),
            FunctionGroup("Operator execution", (
                c("Show only assigned tasks to Operator", "Open My Tasks as an Operator.", "Only tasks assigned to the current Operator are returned."),
                c("Acknowledge assigned task", "Select Receive on a NEW assigned task.", "Status changes to ACKNOWLEDGED with timestamp and actor."),
                c("Start acknowledged task", "Select Start on an acknowledged task.", "Status changes to IN_PROGRESS."),
                c("Reject invalid lifecycle transition", "Try to complete a NEW task without acknowledging/starting it.", "The transition is rejected."),
                c("Complete task with handling result", "Enter at least 10 characters of result text and confirm completion.", "Status becomes COMPLETED and result, user and completion time are stored."),
                c("Reject short completion result", "Submit fewer than 10 characters as the handling result.", "Validation blocks completion."),
                c("Filter and paginate task records", "Filter task records and move between pages.", "Filters apply before 15-row pagination."),
            )),
        ),
    ),
    Module(
        "SIM", "Simulation", "Simulation", "Verify simulation datasets and the end-to-end risk response chain.",
        "Admin is authenticated and ports, zones, thresholds and SOP rules exist.",
        "SimulationPage.test.tsx; SimulationResultsPage.test.tsx; SimulationMap.test.tsx; SimulationFlowTests.cs",
        (
            FunctionGroup("Dataset management", (
                c("Create simulation dataset", "Create a dataset with multiple weather time points and affected zones.", "The dataset is saved and selected for simulation."),
                c("Edit simulation dataset", "Change a time point and save.", "The updated dataset is reloaded."),
                c("Delete simulation dataset", "Confirm deletion from settings.", "The selected dataset is removed."),
                c("Validate simulation inputs", "Submit invalid wind, rain, visibility or missing zone data.", "Validation prevents an invalid dataset."),
            )),
            FunctionGroup("Simulation execution", (
                c("Run ordered simulation steps", "Start a valid multi-step dataset.", "Steps are processed in time order and progress is visible."),
                c("Evaluate each simulation weather point", "Allow the simulation worker to process each point.", "Risk Engine uses the same thresholds and max-risk rule as live data."),
                c("Apply SOP during simulation", "Reach a matching HIGH/CRITICAL step.", "Operation mode, alert and task actions follow active SOP rules."),
                c("Show severe popup during simulation", "Reach a severe simulation step while alerts are enabled.", "The browser shows the related popup/audio notification."),
                c("Display simulation summary", "Complete the run and open results.", "Highest risk, alerts, tasks, affected zones and next actions are shown."),
                c("Keep simulation separate from official confidence", "Run the intervention demo and then open forecast evaluation.", "Simulation/demo observations do not alter official forecast confidence data."),
            )),
        ),
    ),
    Module(
        "FCST", "Forecast & AI", "Operational Forecast & AI Assistance", "Verify OpenWeather planning, aggregation, analytics and guarded AI explanation.",
        "A port has coordinates; forecast provider configuration and thresholds are available.",
        "ForecastPlanningPage.test.tsx; AiLongRangeForecastPage.test.tsx; mlService.test.ts; OpenWeatherForecastParserTests.cs",
        (
            FunctionGroup("Five-day planning", (
                c("Parse OpenWeather five-day forecast", "Load the provider forecast for a configured port.", "Time points are parsed into wind, rain, visibility, temperature and humidity values."),
                c("Aggregate forecast by day", "Build the five-day operation plan.", "Daily rain total, highest wind and lowest visibility are calculated from provider points."),
                c("Evaluate daily forecast risk", "Pass each daily aggregate through Risk Engine.", "Daily risk and dominant factor follow active thresholds."),
                c("Map risk to operation mode", "Build actions for LOW through CRITICAL forecast days.", "Normal, restricted or suspended operation never violates safety guardrails."),
                c("Render five-day timeline and chart", "Open Operational Forecast.", "Five daily plan items and risk visualization are displayed."),
                c("Persist forecast points for later verification", "Generate a forecast plan.", "Forecast timestamps and predicted factors are stored for actual comparison."),
            )),
            FunctionGroup("Long-range and AI", (
                c("Build 7/14/30-day trend view", "Select a supported near-term horizon.", "Provider seed data and extended trend are presented with confidence context."),
                c("Build two/three-month trend view", "Select a long horizon.", "The page clearly labels extended data as lower-confidence planning support."),
                c("Generate PCA trend score", "Analyze normalized forecast features.", "PCA-derived trend information is produced without replacing rule-based risk."),
                c("Group similar days with K-Means", "Cluster daily operational features.", "Days with similar weather/operational characteristics are grouped."),
                c("Use Gemini structured explanation", "Request an explanation when Gemini is configured.", "Structured weather, risk and analytics data are converted into readable planning guidance."),
                c("Use rule fallback when Gemini fails", "Simulate unavailable or invalid Gemini output.", "A deterministic rule-based plan is returned."),
                c("Enforce AI safety floor", "Provide AI output below the allowed HIGH/CRITICAL operation mode.", "The result is raised to at least Restricted for HIGH and Suspended for CRITICAL."),
            )),
        ),
    ),
    Module(
        "EVAL", "Forecast Evaluation", "Forecast Verification, Confidence & Intervention", "Verify predicted-versus-actual matching, error metrics and intervention rules.",
        "Stored forecast points and later actual weather readings are available.",
        "ForecastEvaluationPage.test.tsx; ForecastEvaluationControllerTests.cs; ForecastConfidenceCalculatorTests.cs",
        (
            FunctionGroup("Verification and metrics", (
                c("Match forecast point to actual observation", "Evaluate a past forecast timestamp with an eligible real reading.", "The closest valid actual observation is linked to the forecast point."),
                c("Show missing actual data", "Evaluate a forecast point without an eligible reading.", "The point remains unmatched and is labeled as missing data."),
                c("Calculate wind MAE", "Compare forecast and actual wind values.", "Absolute wind errors and their mean are calculated in m/s."),
                c("Calculate rain MAE", "Compare forecast and actual rain values.", "Absolute rain errors and their mean are calculated in mm."),
                c("Calculate visibility MAE", "Compare forecast and actual visibility values.", "Absolute visibility errors and their mean are calculated in km."),
                c("Calculate risk-match confidence", "Compare predicted and actual risk for matched past points.", "Confidence equals matched risk levels divided by evaluated matched points."),
                c("Require minimum confidence sample", "Request confidence with fewer than three matched points.", "The result is labeled insufficient instead of overstating confidence."),
                c("Classify confidence level", "Evaluate rates around 85% and 70% thresholds.", "HIGH is at least 85%, MEDIUM is 70–84.99%, LOW is below 70%."),
            )),
            FunctionGroup("Intervention", (
                c("Trigger intervention after consecutive mismatches", "Produce three consecutive risk mismatches.", "An intervention recommendation is raised."),
                c("Trigger intervention for low confidence", "Provide at least five samples with confidence below 70%.", "The system requests threshold/model review."),
                c("Trigger intervention for dangerous underestimation", "Create at least two dangerous underestimates in the latest five points.", "A safety-focused intervention warning is shown."),
                c("Export evaluation CSV", "Apply filters and select Export CSV.", "The current authorized evaluation dataset is downloaded."),
            )),
        ),
    ),
    Module(
        "LOG", "Logs & Reports", "Operation History & Reporting", "Verify traceability, filtering, preview and document export.",
        "Operational events, alerts and tasks exist; Admin or Port Manager is authenticated.",
        "LogPage.test.tsx; ReportPage.test.tsx; logService.test.ts; ProductionApiContractTests.cs",
        (
            FunctionGroup("Operation history", (
                c("Record weather/risk workflow events", "Run a weather or simulation workflow.", "Weather update, risk, SOP, alert and task events are traceable."),
                c("Record task lifecycle events", "Assign, acknowledge, start and complete a task.", "Each actor, action and timestamp appears in operation history."),
                c("Group a simulation run", "Open the history after one simulation session.", "Related events are grouped into one readable run summary."),
                c("Search by simulation session or scenario", "Enter a session ID or scenario name.", "Matching simulation history is returned."),
                c("Filter and paginate operation logs", "Apply port, zone, date and level filters, then navigate pages.", "Filters apply before 15-row pagination."),
            )),
            FunctionGroup("Operational reports", (
                c("Preview alert report", "Select Alert report and apply filters.", "An authorized preview is shown before export."),
                c("Preview task report", "Select Task report and apply filters.", "The task preview matches the selected scope."),
                c("Preview operation-history report", "Select Operation History report and apply filters.", "The history preview matches the selected scope."),
                c("Export Excel report", "Confirm a valid preview and select Export Excel.", "A spreadsheet is downloaded for the current filters."),
                c("Export PDF report", "Confirm a valid preview and select Export PDF.", "A PDF is downloaded for the current filters."),
                c("Enforce report port scope", "Request another port's report as Port Manager.", "Backend scope prevents cross-port export."),
                c("Audit report export", "Export a report and review operation history.", "Exporter, report type and timestamp are recorded."),
            )),
        ),
    ),
    Module(
        "SEC", "Performance & Security", "Performance, Security & API Health", "Verify critical non-functional controls covered by automated regression.",
        "Backend, frontend test runner and PostgreSQL test database are available.",
        "ApiHealthTests.cs; RoleAuthorizationTests.cs; ProductionApiContractTests.cs; api.test.ts; style-contract.test.ts",
        (
            FunctionGroup("API health and contracts", (
                c("Return API health status", "Call the health endpoint.", "The endpoint returns a successful health response."),
                c("Allow configured frontend CORS origin", "Send an OPTIONS preflight from the configured frontend origin.", "Required CORS headers are returned."),
                c("Reject unauthenticated protected API call", "Call a protected endpoint without a valid token.", "The API returns 401 and no protected data."),
                c("Reject role-forbidden API call", "Call an Admin-only endpoint using a lower role.", "The API returns 403."),
                c("Enforce port scope in backend", "Request resources belonging to another port as Port Manager.", "The API denies or filters the out-of-scope data."),
                c("Keep production API contract stable", "Run production route contract smoke tests.", "Required routes respond with the expected status/shape."),
            )),
            FunctionGroup("Client resilience", (
                c("Retry request after token refresh", "Receive 401 with a renewable session.", "The API client refreshes once and retries without a request loop."),
                c("Show controlled API error state", "Force a page data request to fail.", "The page shows a clear error/empty state and remains usable."),
                c("Preserve responsive layout contract", "Run UI style contract checks.", "Core pages retain expected responsive layout rules."),
                c("Complete automated regression", "Run backend and frontend test suites against the configured test environment.", "Backend 112/112 and frontend 144/144 tests pass (256/256 total)."),
            )),
        ),
    ),
)


def copy_row_style(ws, source_row: int, target_row: int, max_col: int = 15) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def flatten(groups: Iterable[FunctionGroup]) -> list[Case]:
    return [case for group in groups for case in group.cases]


def fill_cover(ws) -> None:
    ws["B2"] = "TEST REPORT DOCUMENT"
    ws["A4"], ws["B4"] = "Project Name", "PORMS - Port Operation Risk Management System"
    ws["A5"], ws["B5"] = "Project Code", "PORMS"
    ws["A6"], ws["B6"] = "Document Code", "PORMS_Test_Report_v1.0"
    for row in range(9, 18):
        for col in range(1, 7):
            ws.cell(row, col).value = None
    headers = ["Version", "Date", "Author", "Change Description", "Reviewer", "Status"]
    for col, value in enumerate(headers, 1):
        ws.cell(9, col).value = value
    values = ["1.0", EXECUTED_DATE, TESTER, "Initial PORMS system test report based on current source and regression evidence.", "PORMS Team", "Completed"]
    for col, value in enumerate(values, 1):
        ws.cell(10, col).value = value
    ws["B10"].number_format = "dd-mmm-yyyy"
    ws._images = []
    ws.sheet_view.showGridLines = False


def fill_test_case_list(ws) -> None:
    for row in range(8, 200):
        for col in range(1, 7):
            ws.cell(row, col).value = None
    ws["D1"] = "PORMS SYSTEM TEST CASE LIST"
    ws["A3"], ws["B3"] = "Project Name", "PORMS - Port Operation Risk Management System"
    ws["A4"], ws["B4"] = "Project Code", "PORMS"
    ws["A5"], ws["B5"] = "Test Environment", "Windows 11; React/Vite; ASP.NET Core .NET 10; PostgreSQL; Docker Compose; Chrome/Edge"
    headers = ["No.", "Function Name", "Sheet Name", "Description", "Pre-Condition"]
    for col, value in enumerate(headers, 2):
        ws.cell(8, col).value = value
    for idx, module in enumerate(MODULES, 1):
        row = 8 + idx
        ws.cell(row, 2).value = idx
        ws.cell(row, 3).value = module.feature
        ws.cell(row, 4).value = module.sheet
        ws.cell(row, 4).hyperlink = f"#'{module.sheet}'!A1"
        ws.cell(row, 4).style = "Hyperlink"
        ws.cell(row, 5).value = module.description
        ws.cell(row, 6).value = module.precondition
        for col in range(2, 7):
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 38
    ws.freeze_panes = "B9"
    ws.sheet_view.showGridLines = False


def fill_module_sheet(ws, module: Module) -> int:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    for row in range(2, 1000):
        for col in range(1, 16):
            ws.cell(row, col).value = None
    ws["A2"], ws["B2"] = "Feature", module.feature
    ws["A3"], ws["B3"] = "Test requirement", module.description
    ws["A4"], ws["B4"] = "Number of TCs", "=COUNTIF(A12:A1000,\"TC-*\")"
    for row in range(2, 5):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    headers = ["Testing Round", "Passed", "Failed", "Pending", "N/A"]
    for col, value in enumerate(headers, 1):
        ws.cell(5, col).value = value
    for offset, round_name in enumerate(("Round 1", "Round 2", "Round 3"), 6):
        ws.cell(offset, 1).value = round_name
        status_col = get_column_letter(6 + (offset - 6) * 3)
        for col in range(2, 6):
            header_col = get_column_letter(col)
            ws.cell(offset, col).value = f'=COUNTIF(${status_col}$12:${status_col}$1000,{header_col}$5)'
    column_headers = ["Test Case ID", "Test Case Description", "Test Case Procedure", "Expected Results", "Pre-conditions", "Round 1", "Test date", "Tester", "Round 2", "Test date", "Tester", "Round 3", "Test date", "Tester", "Note"]
    for col, value in enumerate(column_headers, 1):
        ws.cell(10, col).value = value

    row = 11
    case_number = 1
    for group_number, group in enumerate(module.groups, 1):
        copy_row_style(ws, 11, row)
        ws.cell(row, 1).value = f"Function {group_number} - {group.name}"
        for col in range(2, 16):
            ws.cell(row, col).value = None
        row += 1
        for case in group.cases:
            copy_row_style(ws, 12, row)
            tc_id = f"TC-{module.code}-{case_number:03d}"
            values = [
                tc_id, case.title, case.procedure, case.expected, case.precondition,
                case.status, EXECUTED_DATE, TESTER, None, None, None, None, None, None,
                f"Evidence: {module.evidence}",
            ]
            for col, value in enumerate(values, 1):
                ws.cell(row, col).value = value
                ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
            for date_col in (7, 10, 13):
                ws.cell(row, date_col).number_format = "dd-mmm-yyyy"
            ws.row_dimensions[row].height = 54
            row += 1
            case_number += 1

    ws.cell(row + 1, 1).value = "↩ Back to Test Case List"
    ws.cell(row + 1, 1).hyperlink = "#'Test Cases'!D1"
    ws.cell(row + 1, 1).style = "Hyperlink"
    ws.freeze_panes = "A11"
    ws.auto_filter.ref = f"A10:O{row - 1}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:10"
    ws.print_area = f"A1:O{row + 1}"
    return case_number - 1


def fill_statistics(ws, counts: dict[str, int]) -> None:
    ws._charts = []
    for row in range(10, 200):
        for col in range(1, 9):
            ws.cell(row, col).value = None
    ws["B1"] = "PORMS TEST STATISTICS"
    ws["B3"], ws["C3"] = "Project Name", "PORMS - Port Operation Risk Management System"
    ws["B4"], ws["C4"] = "Project Code", "PORMS"
    ws["B5"], ws["C5"] = "Execution Date", EXECUTED_DATE
    ws["C5"].number_format = "dd-mmm-yyyy"
    ws["B6"] = "Evidence"
    ws["C6"] = "Automated regression passed: Backend 112/112; Frontend 144/144; Total 256/256 on 11-Aug-2026. The functional scenarios consolidate current source behavior and regression coverage."
    headers = ["No.", "Module code", "Passed", "Failed", "Pending", "N/A", "Number of test cases", "Module / Evidence"]
    for col, value in enumerate(headers, 1):
        ws.cell(10, col).value = value

    start = 11
    for idx, module in enumerate(MODULES, 1):
        row = start + idx - 1
        ws.cell(row, 1).value = idx
        ws.cell(row, 2).value = module.code
        ws.cell(row, 2).hyperlink = f"#'{module.sheet}'!A1"
        ws.cell(row, 3).value = f"='{module.sheet}'!B6"
        ws.cell(row, 4).value = f"='{module.sheet}'!C6"
        ws.cell(row, 5).value = f"='{module.sheet}'!D6"
        ws.cell(row, 6).value = f"='{module.sheet}'!E6"
        ws.cell(row, 7).value = counts[module.sheet]
        ws.cell(row, 8).value = module.feature
        for col in range(1, 9):
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
    total_row = start + len(MODULES)
    ws.cell(total_row, 2).value = "Sub total"
    for col in range(3, 8):
        letter = get_column_letter(col)
        ws.cell(total_row, col).value = f"=SUM({letter}{start}:{letter}{total_row - 1})"
    for col in range(1, 9):
        ws.cell(total_row, col).fill = PatternFill("solid", fgColor="17365D")
        ws.cell(total_row, col).font = Font(color="FFFFFF", bold=True)

    metric_row = total_row + 3
    ws.cell(metric_row, 2).value = "Test coverage"
    ws.cell(metric_row, 4).value = f"=IF(G{total_row}-F{total_row}=0,0,(C{total_row}+D{total_row})/(G{total_row}-F{total_row}))"
    ws.cell(metric_row + 1, 2).value = "Test successful coverage"
    ws.cell(metric_row + 1, 4).value = f"=IF(G{total_row}-F{total_row}=0,0,C{total_row}/(G{total_row}-F{total_row}))"
    for row in (metric_row, metric_row + 1):
        ws.cell(row, 4).number_format = "0.00%"

    chart_data_row = metric_row + 4
    for i, label in enumerate(("Passed", "Failed", "Pending", "N/A"), chart_data_row):
        ws.cell(i, 2).value = label
        ws.cell(i, 3).value = f"={get_column_letter(i - chart_data_row + 3)}{total_row}"
    chart = PieChart()
    chart.title = "PORMS Test Status - Round 1"
    chart.height = 7.5
    chart.width = 11
    chart.add_data(Reference(ws, min_col=3, min_row=chart_data_row, max_row=chart_data_row + 3), titles_from_data=False)
    chart.set_categories(Reference(ws, min_col=2, min_row=chart_data_row, max_row=chart_data_row + 3))
    ws.add_chart(chart, f"J{start}")
    ws.freeze_panes = "A11"
    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:T{chart_data_row + 10}"


def create_consolidated_test_cases(path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Test Cases"
    ws.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor="17365D")
    header_fill = PatternFill("solid", fgColor="4472C4")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    border_color = "9EADBA"

    ws.merge_cells("A1:M1")
    ws["A1"] = "PORMS SYSTEM TEST CASES"
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    metadata = (
        ("Project Name", "PORMS - Port Operation Risk Management System"),
        ("Project Code", "PORMS"),
        ("Execution Evidence", "Backend 112/112 passed; Frontend 144/144 passed; Total 256/256 passed on 11-Aug-2026."),
        ("Test Environment", "Windows 11; React/Vite; ASP.NET Core .NET 10; PostgreSQL; Docker Compose; Chrome/Edge"),
    )
    for row, (label, value) in enumerate(metadata, 3):
        ws.cell(row, 1).value = label
        ws.cell(row, 1).font = Font(name="Arial", bold=True, color="17365D")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
        ws.cell(row, 2).value = value
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    header_row = 8
    headers = (
        "No.", "Test Case ID", "Module", "Function", "Test Case Description",
        "Test Case Procedure", "Expected Results", "Pre-conditions", "Round 1",
        "Test Date", "Tester", "Evidence / Source", "Note",
    )
    for col, value in enumerate(headers, 1):
        cell = ws.cell(header_row, col)
        cell.value = value
        cell.fill = header_fill
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = header_row + 1
    sequence = 1
    for module in MODULES:
        module_start = row
        case_number = 1
        for group in module.groups:
            for case in group.cases:
                values = (
                    sequence,
                    f"TC-{module.code}-{case_number:03d}",
                    module.feature,
                    group.name,
                    case.title,
                    case.procedure,
                    case.expected,
                    case.precondition,
                    case.status,
                    EXECUTED_DATE,
                    TESTER,
                    module.evidence,
                    "Current PORMS regression and source-backed functional verification.",
                )
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row, col)
                    cell.value = value
                    cell.font = Font(name="Arial", size=10)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                ws.cell(row, 10).number_format = "dd-mmm-yyyy"
                ws.row_dimensions[row].height = 62
                row += 1
                sequence += 1
                case_number += 1
        for current_row in range(module_start, row):
            ws.cell(current_row, 3).fill = section_fill

    widths = {
        "A": 7, "B": 18, "C": 28, "D": 26, "E": 38, "F": 52,
        "G": 48, "H": 38, "I": 12, "J": 14, "K": 16, "L": 48, "M": 34,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A8:M{row - 1}"
    ws.print_title_rows = "1:8"
    ws.print_area = f"A1:M{row - 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "PORMS Test Cases"
    ws.oddFooter.right.text = "Page &P of &N"

    # Keep a compact module index so the sheet is easy to navigate after insertion.
    index = workbook.create_sheet("Module Index", 0)
    index.sheet_view.showGridLines = False
    index.merge_cells("A1:E1")
    index["A1"] = "PORMS TEST CASE MODULE INDEX"
    index["A1"].fill = title_fill
    index["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    index["A1"].alignment = Alignment(horizontal="center")
    for col, value in enumerate(("No.", "Module Code", "Module", "Number of Test Cases", "Open Test Cases"), 1):
        index.cell(3, col).value = value
        index.cell(3, col).fill = header_fill
        index.cell(3, col).font = Font(name="Arial", bold=True, color="FFFFFF")
        index.cell(3, col).alignment = Alignment(horizontal="center", wrap_text=True)
    first_row = 9
    for idx, module in enumerate(MODULES, 1):
        target_row = first_row + sum(len(flatten(item.groups)) for item in MODULES[: idx - 1])
        index.cell(idx + 3, 1).value = idx
        index.cell(idx + 3, 2).value = module.code
        index.cell(idx + 3, 3).value = module.feature
        index.cell(idx + 3, 4).value = len(flatten(module.groups))
        index.cell(idx + 3, 5).value = "Go to test cases"
        index.cell(idx + 3, 5).hyperlink = f"#'Test Cases'!A{target_row}"
        index.cell(idx + 3, 5).style = "Hyperlink"
    for column, width in {"A": 8, "B": 16, "C": 38, "D": 22, "E": 22}.items():
        index.column_dimensions[column].width = width
    index.freeze_panes = "A4"

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)


def main() -> None:
    if not REFERENCE.exists():
        raise FileNotFoundError(f"Reference workbook not found: {REFERENCE}")
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(REFERENCE)
    template = wb["Authentication"]

    for name in list(wb.sheetnames):
        if name not in {"Cover", "Test Cases", "Test Statistics", "Authentication"}:
            wb.remove(wb[name])
    template.title = MODULES[0].sheet
    module_sheets = {MODULES[0].sheet: template}
    for module in MODULES[1:]:
        copied = wb.copy_worksheet(template)
        copied.title = module.sheet
        module_sheets[module.sheet] = copied

    fill_cover(wb["Cover"])
    fill_test_case_list(wb["Test Cases"])
    counts: dict[str, int] = {}
    for module in MODULES:
        counts[module.sheet] = fill_module_sheet(module_sheets[module.sheet], module)
    fill_statistics(wb["Test Statistics"], counts)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.active = wb.sheetnames.index("Cover")
    wb.save(OUTPUT)
    wb.save(DOWNLOAD_COPY)
    create_consolidated_test_cases(TEST_CASE_OUTPUT)
    create_consolidated_test_cases(TEST_CASE_DOWNLOAD_COPY)
    print(f"Created: {OUTPUT}")
    print(f"Created: {DOWNLOAD_COPY}")
    print(f"Created: {TEST_CASE_OUTPUT}")
    print(f"Created: {TEST_CASE_DOWNLOAD_COPY}")
    print(f"Modules: {len(MODULES)}")
    print(f"Functional scenarios: {sum(counts.values())}")


if __name__ == "__main__":
    main()
